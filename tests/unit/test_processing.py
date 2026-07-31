"""Tests del job de procesamiento: streaming, validación, particionado y manifiesto."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from common.features import ARCHIVO_MANIFIESTO, RUTA_CURADO
from common.schema import ARCHIVO_FUENTE, ErrorDeEsquema, nombres_columnas
from processing import main as procesamiento
from processing.lector_excel import ErrorDeLectura, leer_encabezado, leer_por_lotes


def _escribir_xlsx(destino: Path, df: pd.DataFrame) -> Path:
    libro = Workbook(write_only=True)
    hoja = libro.create_sheet()
    hoja.append(list(df.columns))
    for fila in df.itertuples(index=False, name=None):
        hoja.append([None if pd.isna(v) else v for v in fila])
    destino.parent.mkdir(parents=True, exist_ok=True)
    libro.save(destino)
    return destino


@pytest.fixture
def raw_local(tmp_path: Path, ss_df, trf_df, pse_df) -> Path:
    """Carpeta con los tres XLSX, tal como los entrega ACH."""
    carpeta = tmp_path / "entrega"
    for fuente, df in (("ss", ss_df), ("trf", trf_df), ("pse", pse_df)):
        _escribir_xlsx(carpeta / ARCHIVO_FUENTE[fuente], df)
    return carpeta


# --------------------------------------------------------------------------- #
# Lector en streaming                                                          #
# --------------------------------------------------------------------------- #
def test_encabezado_se_lee_sin_cargar_el_archivo(raw_local):
    encabezado = leer_encabezado(raw_local / ARCHIVO_FUENTE["pse"])
    assert encabezado == nombres_columnas("pse")


def test_lectura_por_lotes_respeta_el_tamano(raw_local):
    lotes = list(leer_por_lotes(raw_local / ARCHIVO_FUENTE["pse"], tamano_lote=2))
    assert [len(lote) for lote in lotes] == [2, 2]
    assert sum(len(lote) for lote in lotes) == 4


def test_lectura_de_archivo_inexistente_falla_claro(tmp_path):
    with pytest.raises(ErrorDeLectura, match="No existe el archivo"):
        list(leer_por_lotes(tmp_path / "fantasma.xlsx"))


def test_filas_vacias_se_descartan(tmp_path, pse_df):
    ruta = tmp_path / "con_vacias.xlsx"
    libro = Workbook(write_only=True)
    hoja = libro.create_sheet()
    hoja.append(list(pse_df.columns))
    hoja.append([None] * len(pse_df.columns))
    hoja.append(list(pse_df.iloc[0]))
    hoja.append([None] * len(pse_df.columns))
    libro.save(ruta)

    assert sum(len(lote) for lote in leer_por_lotes(ruta)) == 1


# --------------------------------------------------------------------------- #
# Procesamiento completo                                                       #
# --------------------------------------------------------------------------- #
@pytest.fixture
def procesado(storage, raw_local, settings):
    import common.storage as modulo_storage
    from processing.seed import subir

    original = modulo_storage.get_storage
    modulo_storage.get_storage = lambda s=None: storage
    procesamiento.get_storage = lambda s=None: storage
    try:
        import processing.seed as modulo_seed

        modulo_seed.get_storage = lambda s=None: storage
        subir(raw_local, "2026-07-30")
        yield procesamiento.ejecutar(fecha="2026-07-30", settings=settings, tamano_lote=2)
    finally:
        modulo_storage.get_storage = original


def test_el_manifiesto_reporta_filas_columnas_y_hash(procesado):
    assert procesado["filas"] == 10, "2 de SS + 4 de transferencias + 4 de PSE"
    assert procesado["columnas"] == 31 + 26 + 12
    assert len(procesado["hash"]) == 64
    assert procesado["lineage"] == "cedula-v1"
    assert len(procesado["fuentes"]) == 3


def test_el_manifiesto_queda_en_el_bucket(procesado, storage, settings):
    ruta = storage.ruta(settings.bucket_curated, RUTA_CURADO, ARCHIVO_MANIFIESTO)
    assert storage.existe(ruta)
    assert storage.leer_json(ruta)["hash"] == procesado["hash"]


def test_el_dataset_queda_particionado_por_fuente_y_periodo(procesado, storage, settings):
    raiz = storage.ruta(settings.bucket_curated, RUTA_CURADO)
    archivos = storage.listar(raiz, sufijo=".parquet")
    assert archivos, "debe haber parquets escritos"
    assert any("fuente=ss/periodo=2025-01" in a.replace("\\", "/") for a in archivos)
    assert any("fuente=pse/periodo=2025-02" in a.replace("\\", "/") for a in archivos)


def test_cada_fuente_reporta_su_rango_de_periodos(procesado):
    por_fuente = {b["fuente"]: b for b in procesado["fuentes"]}
    assert por_fuente["pse"]["rango_periodos"] == ["2025-01", "2025-02"]
    assert por_fuente["ss"]["filas"] == 2


def test_el_hash_es_estable_entre_corridas(procesado, storage, settings):
    segundo = procesamiento.ejecutar(fecha="2026-07-30", settings=settings, tamano_lote=3)
    assert segundo["hash"] == procesado["hash"], "mismo contenido debe dar el mismo hash"


def test_los_duplicados_exactos_se_descartan(storage, settings, tmp_path, pse_df):
    """Los notebooks originales hacían drop_duplicates(); si no lo replicáramos,
    las métricas de los modelos cambiarían."""

    import processing.seed as modulo_seed

    carpeta = tmp_path / "con_duplicados"
    duplicado = pd.concat([pse_df, pse_df.iloc[[0]]], ignore_index=True)
    for fuente, df in (("ss", None), ("trf", None), ("pse", duplicado)):
        if df is None:
            continue
        _escribir_xlsx(carpeta / ARCHIVO_FUENTE[fuente], df)

    modulo_seed.get_storage = lambda s=None: storage
    procesamiento.get_storage = lambda s=None: storage
    ruta = storage.ruta(settings.bucket_raw, "2026-08-01", ARCHIVO_FUENTE["pse"])
    storage.crear_directorio(ruta.rsplit("/", 1)[0])
    with open(carpeta / ARCHIVO_FUENTE["pse"], "rb") as origen, storage.fs.open(ruta, "wb") as destino:
        destino.write(origen.read())

    manifiesto = procesamiento.ejecutar(fecha="2026-08-01", fuentes=["pse"],
                                        settings=settings, tamano_lote=10)
    bloque = manifiesto["fuentes"][0]
    assert bloque["filas_leidas"] == 5
    assert bloque["filas_duplicadas_descartadas"] == 1
    assert bloque["filas"] == 4


def test_archivo_con_columnas_faltantes_aborta_antes_de_escribir(storage, settings, tmp_path, pse_df):
    import processing.seed as modulo_seed

    modulo_seed.get_storage = lambda s=None: storage
    procesamiento.get_storage = lambda s=None: storage

    incompleto = pse_df.drop(columns=["Valor"])
    local = _escribir_xlsx(tmp_path / "malo" / ARCHIVO_FUENTE["pse"], incompleto)
    remoto = storage.ruta(settings.bucket_raw, "2026-08-02", ARCHIVO_FUENTE["pse"])
    storage.crear_directorio(remoto.rsplit("/", 1)[0])
    with open(local, "rb") as origen, storage.fs.open(remoto, "wb") as destino:
        destino.write(origen.read())

    with pytest.raises(ErrorDeEsquema) as error:
        procesamiento.ejecutar(fecha="2026-08-02", fuentes=["pse"], settings=settings)
    assert "Valor" in str(error.value)
    assert "diccionario de datos" in str(error.value)


def test_carga_inexistente_falla_con_mensaje_util(storage, settings):
    procesamiento.get_storage = lambda s=None: storage
    with pytest.raises(FileNotFoundError, match="Falta el archivo"):
        procesamiento.ejecutar(fecha="1999-01-01", fuentes=["pse"], settings=settings)


def test_el_curado_se_puede_leer_con_la_capa_de_features(procesado, storage):
    from common import features as ft

    tabla = ft.tabla_persona_mes("trf", storage=storage)
    assert tabla["recibido"].iloc[0] == 1_200_000.0

    perfil = ft.perfil_persona_ss(storage=storage).set_index("person_id")
    assert perfil.loc["10000002****", "tipo_persona"] == "Pensionado"
