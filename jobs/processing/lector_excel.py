"""Lectura de los extractos XLSX en streaming.

Los tres archivos rondan el millón de filas y los 100 MB cada uno. Cargarlos
completos con ``pd.read_excel`` —como hacía uno de los scripts originales— reventaba
la memoria del contenedor. Aquí se leen fila a fila con ``openpyxl`` en modo
``read_only`` y se emiten en lotes, de modo que el pico de memoria depende del
tamaño del lote y no del archivo.

Se descartó DuckDB para este paso: no lee XLSX en streaming de forma confiable. Sí
es la herramienta correcta cuando ACH entregue los extractos en CSV o parquet, y en
ese momento este módulo se reemplaza sin tocar el resto del job.
"""

from __future__ import annotations

import logging
from collections.abc import Iterator
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

log = logging.getLogger(__name__)

TAMANO_LOTE = 50_000
LIMITE_FILAS_EXCEL = 1_048_576  # tope de una hoja de cálculo


class ErrorDeLectura(RuntimeError):
    """Problema al leer un archivo de entrada."""


def leer_encabezado(ruta: Path) -> list[str]:
    """Devuelve los nombres de columna sin cargar el archivo completo."""
    libro = load_workbook(ruta, read_only=True, data_only=True)
    try:
        hoja = libro.worksheets[0]
        for fila in hoja.iter_rows(min_row=1, max_row=1, values_only=True):
            return [("" if v is None else str(v)).strip() for v in fila]
        raise ErrorDeLectura(f"{ruta.name} está vacío: no tiene ni siquiera encabezado.")
    finally:
        libro.close()


def leer_por_lotes(ruta: Path, tamano_lote: int = TAMANO_LOTE) -> Iterator[pd.DataFrame]:
    """Emite el contenido del XLSX en DataFrames de ``tamano_lote`` filas.

    Las filas totalmente vacías se descartan: Excel a veces arrastra filas fantasma
    al final de la hoja.
    """
    if not ruta.exists():
        raise ErrorDeLectura(f"No existe el archivo de entrada {ruta}")

    libro = load_workbook(ruta, read_only=True, data_only=True)
    try:
        hoja = libro.worksheets[0]
        iterador = hoja.iter_rows(values_only=True)

        try:
            encabezado_crudo = next(iterador)
        except StopIteration as exc:
            raise ErrorDeLectura(f"{ruta.name} está vacío.") from exc

        columnas = [("" if v is None else str(v)).strip() for v in encabezado_crudo]
        n_columnas = len(columnas)
        acumulado: list[tuple] = []
        total = 0

        for fila in iterador:
            if fila is None or all(v is None for v in fila):
                continue
            if len(fila) != n_columnas:
                fila = tuple(list(fila)[:n_columnas] + [None] * (n_columnas - len(fila)))
            acumulado.append(fila)
            if len(acumulado) >= tamano_lote:
                total += len(acumulado)
                yield pd.DataFrame(acumulado, columns=columnas)
                acumulado = []

        if acumulado:
            total += len(acumulado)
            yield pd.DataFrame(acumulado, columns=columnas)

        log.info("%s: %s filas leídas", ruta.name, f"{total:,}")
        if total >= LIMITE_FILAS_EXCEL - 2:
            log.warning(
                "%s llegó al tope de filas de Excel (%s): el extracto está TRUNCADO. "
                "Los volúmenes son relativos, no censales.",
                ruta.name, f"{LIMITE_FILAS_EXCEL:,}",
            )
    finally:
        libro.close()


def contar_filas(ruta: Path) -> int:
    """Cuenta filas sin materializar el archivo. Solo para diagnóstico."""
    return sum(len(lote) for lote in leer_por_lotes(ruta))
